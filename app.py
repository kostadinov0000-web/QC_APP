import sqlite3
import hashlib
import os
import uuid
import time
from datetime import datetime, timedelta, timezone
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file, send_from_directory
import pandas as pd
import re
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import io
from werkzeug.utils import secure_filename
from openpyxl.drawing.image import Image as XLImage

app = Flask(__name__)

# Production configuration
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your_secret_key_change_in_production')
app.config['DATABASE'] = os.environ.get('DATABASE_PATH', 'quality_control.db')
app.config['DEBUG'] = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'

# Session configuration
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)  # 30 minutes timeout

# Set up logging
if app.config['DEBUG']:
    logging.basicConfig(level=logging.DEBUG)
else:
    logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

DATABASE = app.config['DATABASE']

def get_db_connection():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL
        )
        ''')
        cursor.execute("PRAGMA table_info(users)")
        columns = [info[1] for info in cursor.fetchall()]
        if 'role' not in columns:
            cursor.execute("ALTER TABLE users ADD COLUMN role TEXT DEFAULT 'inspector'")
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_name TEXT NOT NULL,
            drawing_number TEXT NOT NULL,
            drawing_path TEXT,
            comments TEXT,
            UNIQUE(product_name, drawing_number)
        )
        ''')
        # Migration: Check if comments column exists
        cursor.execute("PRAGMA table_info(products)")
        columns = [info[1] for info in cursor.fetchall()]
        if 'comments' not in columns:
            cursor.execute("ALTER TABLE products ADD COLUMN comments TEXT")
        
        # Add molds table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS molds (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            mold_name TEXT NOT NULL,
            mold_number TEXT NOT NULL,
            total_cycles INTEGER DEFAULT 0,
            maintenance_threshold INTEGER DEFAULT 50000,
            last_maintenance_date TEXT,
            status TEXT DEFAULT 'active',
            created_date TEXT NOT NULL,
            specifications_pdf TEXT,
            FOREIGN KEY (product_id) REFERENCES products(id),
            UNIQUE(product_id, mold_number)
        )
        ''')
        
        # Add rework_history table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS rework_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mold_id INTEGER NOT NULL,
            rework_type TEXT NOT NULL,
            rework_date TEXT NOT NULL,
            technician TEXT NOT NULL,
            description TEXT,
            before_photo_path TEXT,
            after_photo_path TEXT,
            parts_replaced TEXT,
            cost REAL,
            FOREIGN KEY (mold_id) REFERENCES molds(id)
        )
        ''')
        
        # Add maintenance_schedule table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS maintenance_schedule (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mold_id INTEGER NOT NULL,
            maintenance_type TEXT NOT NULL,
            scheduled_date TEXT NOT NULL,
            completed_date TEXT,
            technician TEXT,
            checklist_items TEXT,
            notes TEXT,
            technician_notes TEXT,
            status TEXT DEFAULT 'scheduled',
            FOREIGN KEY (mold_id) REFERENCES molds(id)
        )
        ''')
        
        # Migration: Add technician_notes column if it doesn't exist
        cursor.execute("PRAGMA table_info(maintenance_schedule)")
        maintenance_columns = [info[1] for info in cursor.fetchall()]
        if 'technician_notes' not in maintenance_columns:
            cursor.execute("ALTER TABLE maintenance_schedule ADD COLUMN technician_notes TEXT")
        
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS dimensions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            dimension_name TEXT NOT NULL,
            nominal_value REAL NOT NULL,
            tolerance_minus REAL NOT NULL,
            tolerance_plus REAL NOT NULL,
            FOREIGN KEY (product_id) REFERENCES products(id),
            UNIQUE(product_id, dimension_name)
        )
        ''')
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS measurements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            dimension_id INTEGER NOT NULL,
            measured_value REAL NOT NULL,
            measurement_date TEXT NOT NULL,
            machine_number TEXT,
            count INTEGER,
            inspector TEXT,
            FOREIGN KEY (product_id) REFERENCES products(id),
            FOREIGN KEY (dimension_id) REFERENCES dimensions(id)
        )
        ''')
        # Migration: Check if batch_number exists and rename to machine_number, add count, and add shift
        cursor.execute("PRAGMA table_info(measurements)")
        columns = [info[1] for info in cursor.fetchall()]
        if 'batch_number' in columns and 'machine_number' not in columns:
            cursor.execute("ALTER TABLE measurements RENAME COLUMN batch_number TO machine_number")
        if 'count' not in columns:
            cursor.execute("ALTER TABLE measurements ADD COLUMN count INTEGER")
        if 'shift' not in columns:
            cursor.execute("ALTER TABLE measurements ADD COLUMN shift TEXT")
        if 'submission_id' not in columns:
            cursor.execute("ALTER TABLE measurements ADD COLUMN submission_id TEXT")
        cursor.execute("SELECT COUNT(*) FROM users WHERE username='admin'")
        if cursor.fetchone()[0] == 0:
            # Use environment variable for admin password or generate a secure default
            admin_password = os.environ.get('ADMIN_PASSWORD', 'admin123')
            password_hash = hashlib.sha256(admin_password.encode()).hexdigest()
            cursor.execute("INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
                          ("admin", password_hash, "admin"))
            logger.info("Default admin user created. Please change the password in production.")
        
        # Add machine_last_product table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS machine_last_product (
            machine_number TEXT PRIMARY KEY,
            last_product_id INTEGER,
            last_count INTEGER,
            last_measurement_id INTEGER,
            last_update TEXT
        )
        ''')
        
        # Add machine_mold_assignments table for tracking current mold assignments
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS machine_mold_assignments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            machine_number TEXT NOT NULL,
            mold_id INTEGER NOT NULL,
            assigned_date TEXT NOT NULL,
            assigned_by TEXT NOT NULL,
            status TEXT DEFAULT 'active',
            FOREIGN KEY (mold_id) REFERENCES molds(id)
        )
        ''')
        
        # Add mold_problems table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS mold_problems (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mold_id INTEGER NOT NULL,
            problem_type TEXT NOT NULL,
            description TEXT,
            inspector TEXT NOT NULL,
            report_date TEXT NOT NULL,
            comments TEXT,
            status TEXT DEFAULT 'open',
            FOREIGN KEY (mold_id) REFERENCES molds(id)
        )
        ''')
        # Migration: Check if drawing_path_2 column exists
        cursor.execute("PRAGMA table_info(products)")
        columns = [info[1] for info in cursor.fetchall()]
        if 'drawing_path_2' not in columns:
            cursor.execute("ALTER TABLE products ADD COLUMN drawing_path_2 TEXT")
        
        # Migration: Check if specifications_pdf column exists in molds table
        cursor.execute("PRAGMA table_info(molds)")
        mold_columns = [info[1] for info in cursor.fetchall()]
        if 'specifications_pdf' not in mold_columns:
            cursor.execute("ALTER TABLE molds ADD COLUMN specifications_pdf TEXT")
        
        # Migration: Add comment column to molds if it doesn't exist
        cursor.execute("PRAGMA table_info(molds)")
        mold_columns = [info[1] for info in cursor.fetchall()]
        if 'comment' not in mold_columns:
            cursor.execute("ALTER TABLE molds ADD COLUMN comment TEXT")
        
        # Migration: Add completed_date column to rework_history if it doesn't exist
        cursor.execute("PRAGMA table_info(rework_history)")
        rework_columns = [info[1] for info in cursor.fetchall()]
        if 'completed_date' not in rework_columns:
            cursor.execute("ALTER TABLE rework_history ADD COLUMN completed_date TEXT")
        
        conn.commit()

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def convert_to_iso_date(local_date):
    # Try ISO format first (YYYY-MM-DD) - from HTML5 date inputs
    try:
        return datetime.strptime(local_date, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        pass
    
    # Try local format (DD-MM-YYYY) - legacy format
    try:
        return datetime.strptime(local_date, "%d-%m-%Y").strftime("%Y-%m-%d")
    except ValueError:
        raise ValueError("Invalid date format. Use YYYY-MM-DD or DD-MM-YYYY.")

def get_bulgarian_time():
    """Get current time in Bulgarian timezone"""
    # Server is in Amsterdam (UTC+1), Bulgaria is UTC+2, so we need +1 hour from server time
    # During DST: Amsterdam UTC+2, Bulgaria UTC+3, still +1 hour difference
    return datetime.now() + timedelta(hours=1)

def get_bulgarian_time_string():
    """Get current Bulgarian time as string for database storage"""
    return get_bulgarian_time().strftime("%Y-%m-%d %H:%M:%S")

def convert_to_local_date(iso_date):
    try:
        # Try to parse with time first
        dt = datetime.strptime(iso_date, "%Y-%m-%d %H:%M:%S")
        return dt.strftime("%d-%m-%Y %H:%M")
    except ValueError:
        try:
            # Fallback to date only
            return datetime.strptime(iso_date, "%Y-%m-%d").strftime("%d-%m-%Y")
        except ValueError:
            return iso_date

def check_session_timeout():
    """Check if session has expired due to inactivity"""
    if 'user' in session:
        # Check if last_activity exists and if session has expired
        if 'last_activity' in session:
            last_activity = datetime.fromisoformat(session['last_activity'])
            if get_bulgarian_time() - last_activity > app.config['PERMANENT_SESSION_LIFETIME']:
                session.clear()
                flash('Your session has expired due to inactivity. Please log in again.', 'warning')
                return True
        
        # Update last activity time
        session['last_activity'] = get_bulgarian_time().isoformat()
        session.permanent = True
    
    return False

@app.before_request
def require_login():
    # Skip timeout check for static files, login page, and session status
    if request.endpoint in ['login', 'static', 'serve_drawing', 'get_drawings', 'serve_pdfjs', 'session_status']:
        return
    
    # Check session timeout
    if check_session_timeout():
        return redirect(url_for('login'))
    
    # Check if user is logged in
    if request.endpoint not in ['login', 'static', 'get_product', 'serve_drawing', 'get_drawings', 'serve_pdfjs', 'debug_pdf', 'debug_pdf_viewer', 'upload_drawing', 'add_drawing_to_product', 'view_tolerance_table', 'session_status'] and 'user' not in session:
        return redirect(url_for('login'))

@app.route('/')
def index():
    if 'user' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if not username or not password:
            flash('Потребителското име и паролата са задължителни', 'error')
            return render_template('login.html')
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT password_hash, role FROM users WHERE username=?", (username,))
            user = cursor.fetchone()
            if user and user['password_hash'] == hash_password(password):
                session['user'] = username
                session['role'] = user['role']
                session['last_activity'] = get_bulgarian_time().isoformat()
                session.permanent = True
                return redirect(url_for('dashboard'))
            flash('Невалидно потребителско име или парола', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out successfully.', 'info')
    return redirect(url_for('login'))

@app.route('/session_status')
def session_status():
    """Check session status via AJAX"""
    if 'user' not in session:
        return jsonify({'status': 'expired', 'message': 'Session expired'})
    
    # Check if session has expired
    if 'last_activity' in session:
        last_activity = datetime.fromisoformat(session['last_activity'])
        time_left = app.config['PERMANENT_SESSION_LIFETIME'] - (get_bulgarian_time() - last_activity)
        
        if time_left.total_seconds() <= 0:
            session.clear()
            return jsonify({'status': 'expired', 'message': 'Session expired due to inactivity'})
        
        return jsonify({
            'status': 'active',
            'time_left_minutes': int(time_left.total_seconds() / 60),
            'warning': time_left.total_seconds() <= 300  # Warning if less than 5 minutes
        })
    
    return jsonify({'status': 'active', 'time_left_minutes': 30})

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html', user=session.get('user'), role=session.get('role'))

@app.route('/products', methods=['GET', 'POST'])
def products():
    # Check if user is admin
    if session.get('role') != 'admin':
        flash('Достъп отказан. Необходими са администраторски права.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get pagination and search parameters
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    per_page = 50  # Products per page
    
    with get_db_connection() as conn:
        cursor = conn.cursor()
        
        # Build search query
        if search:
            search_pattern = f"%{search}%"
            count_query = "SELECT COUNT(*) FROM products WHERE product_name LIKE ? OR drawing_number LIKE ? OR comments LIKE ?"
            products_query = """SELECT id, product_name, drawing_number, drawing_path, drawing_path_2, comments 
                               FROM products 
                               WHERE product_name LIKE ? OR drawing_number LIKE ? OR comments LIKE ?
                               ORDER BY product_name 
                               LIMIT ? OFFSET ?"""
            cursor.execute(count_query, (search_pattern, search_pattern, search_pattern))
            total = cursor.fetchone()[0]
            
            offset = (page - 1) * per_page
            cursor.execute(products_query, (search_pattern, search_pattern, search_pattern, per_page, offset))
            products = cursor.fetchall()
        else:
            cursor.execute("SELECT COUNT(*) FROM products")
            total = cursor.fetchone()[0]
            
            offset = (page - 1) * per_page
            cursor.execute("SELECT id, product_name, drawing_number, drawing_path, drawing_path_2, comments FROM products ORDER BY product_name LIMIT ? OFFSET ?", (per_page, offset))
            products = cursor.fetchall()
        
        # Calculate pagination info
        total_pages = (total + per_page - 1) // per_page
        has_prev = page > 1
        has_next = page < total_pages
        if request.method == 'POST':
            action = request.form.get('action')

            if action == 'add_product':
                try:
                    product_name = request.form['product_name'].strip()
                    drawing_number = request.form['drawing_number'].strip()
                except KeyError as e:
                    logger.error(f"Missing form field: {e}")
                    flash(f'Липсва задължително поле: {e}', 'error')
                    return render_template('products.html', 
                                         products=products, 
                                         role=session.get('role'),
                                         page=page,
                                         total_pages=total_pages,
                                         has_prev=has_prev,
                                         has_next=has_next,
                                         search=search,
                                         total=total)
                drawing_path = request.form.get('drawing_path', '')
                drawing_path_2 = request.form.get('drawing_path_2', '')
                if not product_name or not drawing_number:
                    flash('Името на продукта и номерът на чертежа са задължителни', 'error')
                else:
                    # Check for case-insensitive duplicates before inserting (with whitespace trimming)
                    cursor.execute(
                        "SELECT id, product_name, drawing_number FROM products WHERE LOWER(TRIM(product_name)) = LOWER(TRIM(?)) AND LOWER(TRIM(drawing_number)) = LOWER(TRIM(?))",
                        (product_name, drawing_number)
                    )
                    existing_product = cursor.fetchone()
                    
                    if existing_product:
                        error_msg = f'Дублиран продукт! Продукт с това име и номер на чертеж вече съществува: "{existing_product["product_name"]}" с номер "{existing_product["drawing_number"]}"'
                        flash(error_msg, 'error')
                        logger.warning(f"Duplicate product attempt blocked: {product_name} with drawing {drawing_number} (existing: {existing_product['product_name']} with {existing_product['drawing_number']})")
                    else:
                        # Normalize drawing_path
                        if drawing_path and not drawing_path.startswith('http'):
                            drawing_path = f'drawings/{drawing_path}' if not drawing_path.startswith('drawings/') else drawing_path
                        if drawing_path_2 and not drawing_path_2.startswith('http'):
                            drawing_path_2 = f'drawings/{drawing_path_2}' if not drawing_path_2.startswith('drawings/') else drawing_path_2
                        try:
                            cursor.execute(
                                "INSERT INTO products (product_name, drawing_number, drawing_path, drawing_path_2) VALUES (?, ?, ?, ?)",
                                (product_name, drawing_number, drawing_path, drawing_path_2)
                            )
                            product_id = cursor.lastrowid
                            # Automatically create a mold for the new product
                            mold_name = f"Mold for {product_name}"
                            mold_number = f"M{product_id:04d}"
                            current_date = get_bulgarian_time_string()
                            cursor.execute(
                                "INSERT INTO molds (product_id, mold_name, mold_number, created_date) VALUES (?, ?, ?, ?)",
                                (product_id, mold_name, mold_number, current_date)
                            )
                            conn.commit()
                            flash('Продуктът и неговата матрица са добавени успешно', 'success')
                            logger.info(f"New product created: {product_name} with drawing number: {drawing_number}")
                        except sqlite3.IntegrityError:
                            flash('Продукт с това име и номер на чертеж вече съществува', 'error')
            elif action == 'delete_product':
                product_id = request.form['product_id']
                cursor.execute("DELETE FROM measurements WHERE dimension_id IN (SELECT id FROM dimensions WHERE product_id=?)", (product_id,))
                cursor.execute("DELETE FROM dimensions WHERE product_id=?", (product_id,))
                cursor.execute("DELETE FROM products WHERE id=?", (product_id,))
                conn.commit()
                flash('Продуктът е изтрит успешно', 'success')
                
        # After POST operations, recalculate products with pagination
        if request.method == 'POST':
            # Build search query again after POST
            if search:
                search_pattern = f"%{search}%"
                count_query = "SELECT COUNT(*) FROM products WHERE product_name LIKE ? OR drawing_number LIKE ? OR comments LIKE ?"
                products_query = """SELECT id, product_name, drawing_number, drawing_path, drawing_path_2, comments 
                                   FROM products 
                                   WHERE product_name LIKE ? OR drawing_number LIKE ? OR comments LIKE ?
                                   ORDER BY product_name 
                                   LIMIT ? OFFSET ?"""
                cursor.execute(count_query, (search_pattern, search_pattern, search_pattern))
                total = cursor.fetchone()[0]
                
                offset = (page - 1) * per_page
                cursor.execute(products_query, (search_pattern, search_pattern, search_pattern, per_page, offset))
                products = cursor.fetchall()
            else:
                cursor.execute("SELECT COUNT(*) FROM products")
                total = cursor.fetchone()[0]
                
                offset = (page - 1) * per_page
                cursor.execute("SELECT id, product_name, drawing_number, drawing_path, drawing_path_2, comments FROM products ORDER BY product_name LIMIT ? OFFSET ?", (per_page, offset))
                products = cursor.fetchall()
            
            # Recalculate pagination info
            total_pages = (total + per_page - 1) // per_page
            has_prev = page > 1
            has_next = page < total_pages
            
    return render_template('products.html', 
                         products=products, 
                         role=session.get('role'),
                         page=page,
                         total_pages=total_pages,
                         has_prev=has_prev,
                         has_next=has_next,
                         search=search,
                         total=total)

@app.route('/get_dimensions/<int:product_id>')
def get_dimensions(product_id):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
        SELECT id, dimension_name, nominal_value, tolerance_minus, tolerance_plus 
        FROM dimensions 
        WHERE product_id=?
        ''', (product_id,))
        dimensions = [dict(row) for row in cursor.fetchall()]
    return jsonify(dimensions)

@app.route('/get_product/<int:product_id>')
def get_product(product_id):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id, product_name, drawing_number, drawing_path, drawing_path_2, comments FROM products WHERE id=?", (product_id,))
        product = cursor.fetchone()
        if product:
            return jsonify(dict(product))
        return jsonify({'error': 'Продуктът не е намерен'}), 404

@app.route('/drawings/<filename>')
def serve_drawing(filename):
    logger.debug(f"Attempting to serve drawing: {filename}")
    file_path = os.path.join('static', 'drawings', filename)
    if not os.path.exists(file_path):
        logger.error(f"File not found: {file_path}")
        return jsonify({'error': 'Файлът не е намерен'}), 404
    if not filename.lower().endswith('.pdf'):
        logger.error(f"Invalid file type: {filename}")
        return jsonify({'error': 'Невалиден тип файл'}), 400
    logger.debug(f"Serving PDF: {file_path}")
    # Serve PDF directly - much simpler and more reliable
    return send_file(file_path, mimetype='application/pdf')

@app.route('/get_drawings')
def get_drawings():
    # Check if user is admin
    if session.get('role') != 'admin':
        return jsonify({'error': 'Достъп отказан. Необходими са администраторски права.'}), 403
    
    drawings_dir = os.path.join('static', 'drawings')
    try:
        pdf_files = [f for f in os.listdir(drawings_dir) if f.lower().endswith('.pdf')]
        logger.debug(f"Found PDF files: {pdf_files}")
        return jsonify(pdf_files)
    except FileNotFoundError:
        logger.error(f"Drawings directory not found: {drawings_dir}")
        return jsonify([])

@app.route('/pdfjs/<path:filename>')
def serve_pdfjs(filename):
    logger.debug(f"Serving PDF.js file: {filename}")
    response = send_from_directory('static/pdfjs-4.7.76/web', filename)
    # Add CORS headers for PDF.js
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response

@app.route('/add_dimension', methods=['POST'])
def add_dimension():
    # Check if user is admin
    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Достъп отказан. Необходими са администраторски права.'})
    
    product_id = request.form['product_id']
    dimension_name = request.form['dimension_name']
    nominal_value = request.form['nominal_value']
    tolerance_minus = request.form['tolerance_minus']
    tolerance_plus = request.form['tolerance_plus']
    if not all([product_id, dimension_name, nominal_value, tolerance_minus, tolerance_plus]):
        return jsonify({'status': 'error', 'message': 'Всички полета са задължителни'})
    try:
        nominal_value = float(nominal_value)
        tolerance_minus = float(tolerance_minus)
        tolerance_plus = float(tolerance_plus)
    except ValueError:
        return jsonify({'status': 'error', 'message': 'Номиналните стойности и толерансите трябва да бъдат числа'})
    with get_db_connection() as conn:
        cursor = conn.cursor()
        try:
            cursor.execute(
                "INSERT INTO dimensions (product_id, dimension_name, nominal_value, tolerance_minus, tolerance_plus) VALUES (?, ?, ?, ?, ?)",
                (product_id, dimension_name, nominal_value, tolerance_minus, tolerance_plus)
            )
            conn.commit()
            return jsonify({'status': 'success', 'message': 'Размерът е добавен успешно'})
        except sqlite3.IntegrityError:
            return jsonify({'status': 'error', 'message': 'Име на размер вече съществува за този продукт'})
        except sqlite3.OperationalError as e:
            return jsonify({'status': 'error', 'message': f'Неуспешно добавяне на размер: {e}'})

@app.route('/update_dimension', methods=['POST'])
def update_dimension():
    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Достъп отказан. Необходими са администраторски права.'})
    dimension_id = request.form['dimension_id']
    dimension_name = request.form['dimension_name']
    nominal_value = request.form['nominal_value']
    tolerance_minus = request.form['tolerance_minus']
    tolerance_plus = request.form['tolerance_plus']
    
    try:
        nominal_value = float(nominal_value)
        tolerance_minus = float(tolerance_minus)
        tolerance_plus = float(tolerance_plus)
    except ValueError:
        return jsonify({'status': 'error', 'message': 'Невалидни числови стойности'})
    
    with get_db_connection() as conn:
        cursor = conn.cursor()
        try:
            cursor.execute(
                "UPDATE dimensions SET dimension_name=?, nominal_value=?, tolerance_minus=?, tolerance_plus=? WHERE id=?",
                (dimension_name, nominal_value, tolerance_minus, tolerance_plus, dimension_id)
            )
            conn.commit()
            return jsonify({'status': 'success', 'message': 'Размерът е актуализиран успешно'})
        except Exception as e:
            return jsonify({'status': 'error', 'message': f'Неуспешно актуализиране на размер: {e}'})

@app.route('/delete_dimension', methods=['POST'])
def delete_dimension():
    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Достъп отказан. Необходими са администраторски права.'})
    
    dimension_id = request.form['dimension_id']
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM measurements WHERE dimension_id=?", (dimension_id,))
        cursor.execute("DELETE FROM dimensions WHERE id=?", (dimension_id,))
        conn.commit()
    return jsonify({'status': 'success', 'message': 'Размерът е изтрит успешно'})

@app.route('/update_product_comments', methods=['POST'])
def update_product_comments():
    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Достъп отказан. Необходими са администраторски права.'})
    
    product_id = request.form['product_id']
    comments = request.form['comments']
    
    with get_db_connection() as conn:
        cursor = conn.cursor()
        try:
            cursor.execute("UPDATE products SET comments=? WHERE id=?", (comments, product_id))
            conn.commit()
            return jsonify({'status': 'success', 'message': 'Коментарите са актуализирани успешно'})
        except Exception as e:
            return jsonify({'status': 'error', 'message': f'Неуспешно актуализиране на коментари: {e}'})

@app.route('/measurements', methods=['GET', 'POST'])
def measurements():
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id, product_name, drawing_path, drawing_path_2, comments FROM products")
        products = [dict(row) for row in cursor.fetchall()]
        current_date = get_bulgarian_time().strftime('%d-%m-%Y')
        if request.method == 'POST':
            product_id = int(request.form['product_id'])
            machine_number = request.form['machine_number']
            count = int(request.form['count'])
            measurement_date = request.form['measurement_date']
            shift = request.form.get('shift', '')
            if not re.match(r"\d{2}-\d{2}-\d{4}", measurement_date):
                flash('Невалиден формат на датата (използвайте ДД-ММ-ГГГГ)', 'error')
                return render_template('measurements.html', products=products, role=session.get('role'), current_date=current_date)
            try:
                iso_date = datetime.strptime(measurement_date, "%d-%m-%Y").strftime("%Y-%m-%d")
                iso_date = f"{iso_date} {get_bulgarian_time().strftime('%H:%M:%S')}"
            except ValueError:
                flash('Невалиден формат на датата (използвайте ДД-ММ-ГГГГ)', 'error')
                return render_template('measurements.html', products=products, role=session.get('role'), current_date=current_date)
            if count <= 0:
                flash('Броят трябва да бъде положително цяло число', 'error')
                return render_template('measurements.html', products=products, role=session.get('role'), current_date=current_date)
            if not machine_number:
                flash('Номерът на машината е задължителен', 'error')
                return render_template('measurements.html', products=products, role=session.get('role'), current_date=current_date)
            measurements = []
            for key in request.form:
                if key.startswith('measured_value_'):
                    dimension_id = key.split('_')[-1]
                    measured_value = request.form[key]
                    if measured_value:
                        try:
                            measured_value = float(measured_value)
                            measurements.append((product_id, dimension_id, measured_value, iso_date, machine_number, count, session['user'], shift))
                        except ValueError:
                            flash(f'Невалидна стойност на измерване за размер ID {dimension_id}', 'error')
                            return render_template('measurements.html', products=products, role=session.get('role'), current_date=current_date)
            if not measurements:
                flash('Няма въведени измервания', 'error')
            else:
                # Generate unique submission ID to prevent duplicates
                submission_id = f"{product_id}_{machine_number}_{iso_date.replace(' ', '_').replace(':', '')}_{uuid.uuid4().hex[:8]}"
                
                # Check if measurements with similar parameters already exist (duplicate prevention)
                cursor.execute("""
                    SELECT COUNT(*) FROM measurements 
                    WHERE product_id = ? AND machine_number = ? AND date(measurement_date) = date(?) AND inspector = ?
                """, (product_id, machine_number, iso_date, session['user']))
                existing_count = cursor.fetchone()[0]
                
                if existing_count > 0:
                    # Check if this might be a duplicate submission
                    cursor.execute("""
                        SELECT measurement_date FROM measurements 
                        WHERE product_id = ? AND machine_number = ? AND date(measurement_date) = date(?) AND inspector = ?
                        ORDER BY measurement_date DESC LIMIT 1
                    """, (product_id, machine_number, iso_date, session['user']))
                    last_measurement = cursor.fetchone()
                    
                    if last_measurement:
                        last_time = datetime.strptime(last_measurement[0], "%Y-%m-%d %H:%M:%S")
                        current_time = datetime.strptime(iso_date, "%Y-%m-%d %H:%M:%S")
                        time_diff = (current_time - last_time).total_seconds()
                        
                        # If last measurement was within 5 minutes, likely a duplicate
                        if time_diff < 300:  # 5 minutes
                            flash('Възможно дублиране! Измервания за този продукт, машина и дата вече са записани преди малко. Моля, проверете дали не се опитвате да запишете същите данни отново.', 'warning')
                            return render_template('measurements.html', products=products, role=session.get('role'), current_date=current_date)
                
                # Add submission_id to each measurement
                measurements_with_id = []
                for measurement in measurements:
                    measurements_with_id.append(measurement + (submission_id,))
                
                # 1. Check last product for this machine
                cursor.execute("SELECT last_product_id, last_count FROM machine_last_product WHERE machine_number=?", (machine_number,))
                last_row = cursor.fetchone()
                if last_row and last_row['last_product_id'] is not None and last_row['last_product_id'] != product_id:
                    # Add last_count to the previous product's mold cycle count
                    cursor.execute("UPDATE molds SET total_cycles = total_cycles + ? WHERE product_id = ?", (last_row['last_count'], last_row['last_product_id']))
                # 2. Insert measurements with submission_id
                cursor.executemany(
                    "INSERT INTO measurements (product_id, dimension_id, measured_value, measurement_date, machine_number, count, inspector, shift, submission_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    measurements_with_id
                )
                # 3. Update machine_last_product
                # Get the last inserted measurement id
                cursor.execute("SELECT id FROM measurements WHERE product_id=? AND machine_number=? ORDER BY id DESC LIMIT 1", (product_id, machine_number))
                last_measurement_id = cursor.fetchone()
                last_measurement_id = last_measurement_id['id'] if last_measurement_id else None
                cursor.execute("REPLACE INTO machine_last_product (machine_number, last_product_id, last_count, last_measurement_id, last_update) VALUES (?, ?, ?, ?, ?)",
                    (machine_number, product_id, count, last_measurement_id, iso_date))
                
                # Update machine-mold assignments
                # First, get the mold for this product
                cursor.execute("SELECT id FROM molds WHERE product_id = ? LIMIT 1", (product_id,))
                mold_result = cursor.fetchone()
                if mold_result:
                    mold_id = mold_result['id']
                    current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    inspector = session.get('user', 'Unknown')
                    
                    # Check if there's already an active assignment for this machine
                    cursor.execute(
                        "SELECT id FROM machine_mold_assignments WHERE machine_number = ? AND status = 'active'",
                        (machine_number,)
                    )
                    existing_assignment = cursor.fetchone()
                    
                    if existing_assignment:
                        # Update existing assignment
                        cursor.execute(
                            "UPDATE machine_mold_assignments SET mold_id = ?, assigned_date = ?, assigned_by = ? WHERE id = ?",
                            (mold_id, current_date, inspector, existing_assignment['id'])
                        )
                    else:
                        # Create new assignment
                        cursor.execute(
                            "INSERT INTO machine_mold_assignments (machine_number, mold_id, assigned_date, assigned_by) VALUES (?, ?, ?, ?)",
                            (machine_number, mold_id, current_date, inspector)
                        )
                
                conn.commit()
                flash(f'{len(measurements)} измервания са запазени успешно', 'success')
    return render_template('measurements.html', products=products, role=session.get('role'), current_date=current_date)

@app.route('/reports', methods=['GET', 'POST'])
def reports():
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id, product_name FROM products")
        products = [dict(row) for row in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT machine_number FROM measurements")
        machines = [row['machine_number'] for row in cursor.fetchall() if row['machine_number']]
        report_data = []
        headers = []
        end_date = datetime.now().strftime('%d-%m-%Y')
        start_date = (datetime.now() - timedelta(days=30)).strftime('%d-%m-%Y')
        if request.method == 'POST':
            report_type = request.form.get('report_type', 'standard')
            product_id = request.form.get('product_id')
            start_date = request.form['start_date']
            end_date = request.form['end_date']
            if not product_id:
                flash('Моля, изберете продукт, преди да генерирате справка.', 'error')
                return render_template('reports.html', products=products, machines=machines, report_data=report_data, headers=headers, role=session.get('role'), start_date=start_date, end_date=end_date)
            if not re.match(r"\d{2}-\d{2}-\d{4}", start_date) or not re.match(r"\d{2}-\d{2}-\d{4}", end_date):
                flash('Невалиден формат на датата (използвайте ДД-ММ-ГГГГ)', 'error')
                return render_template('reports.html', products=products, machines=machines, report_data=report_data, headers=headers, role=session.get('role'), start_date=start_date, end_date=end_date)
            try:
                iso_start_date = convert_to_iso_date(start_date)
                iso_end_date = convert_to_iso_date(end_date)
            except ValueError:
                flash('Невалиден формат на датата (използвайте ДД-ММ-ГГГГ)', 'error')
                return render_template('reports.html', products=products, machines=machines, report_data=report_data, headers=headers, role=session.get('role'), start_date=start_date, end_date=end_date)
            
            query = '''
            SELECT p.product_name, d.dimension_name, m.measured_value, d.nominal_value, d.tolerance_plus, d.tolerance_minus,
                   m.measurement_date, m.inspector, m.machine_number, m.count, m.shift
            FROM measurements m
            JOIN dimensions d ON m.dimension_id = d.id
            JOIN products p ON m.product_id = p.id
            WHERE date(m.measurement_date) BETWEEN ? AND ?
            {} 
            ORDER BY p.product_name, d.dimension_name, m.measurement_date DESC
            '''
            conditions = []
            params = [iso_start_date, iso_end_date]
            if product_id:
                conditions.append("m.product_id = ?")
                params.append(product_id)
            query = query.format("AND " + " AND ".join(conditions) if conditions else "")
            cursor.execute(query, params)
            report_data = [dict(row) for row in cursor.fetchall()]
            for row in report_data:
                row['measurement_date'] = convert_to_local_date(row['measurement_date'])
            if report_type == 'detailed':
                # Add in_tolerance field for each row
                for row in report_data:
                    nominal = row['nominal_value']
                    minus = row['tolerance_minus']
                    plus = row['tolerance_plus']
                    measured = row['measured_value']
                    row['in_tolerance'] = (nominal - minus) <= measured <= (nominal + plus)
                headers = ["Product", "Dimension", "Measured Value", "Nominal", "Tolerance (+/-)", "Measurement Date", "Inspector", "Machine", "Count", "Shift", "Tol. check"]
            else:
                headers = ["Product", "Dimension", "Measured Value", "Nominal", "Tolerance (+/-)", "Measurement Date", "Inspector", "Machine", "Count", "Shift"]
        return render_template('reports.html', products=products, machines=machines, report_data=report_data, headers=headers, role=session.get('role'), start_date=start_date, end_date=end_date)

@app.route('/export_excel', methods=['POST'])
def export_excel():
    if not session.get('user'):
        return redirect(url_for('login'))
    
    with get_db_connection() as conn:
        cursor = conn.cursor()
        
        # Get form data
        report_type = 'detailed_measurements'  # Always use detailed measurements
        product_id = request.form.get('product_id')
        start_date = request.form['start_date']
        end_date = request.form['end_date']
        
        # Validate date format
        if not re.match(r"\d{2}-\d{2}-\d{4}", start_date) or not re.match(r"\d{2}-\d{2}-\d{4}", end_date):
            flash('Невалиден формат на датата (използвайте ДД-ММ-ГГГГ)', 'error')
            return redirect(url_for('reports'))
        
        try:
            iso_start_date = convert_to_iso_date(start_date)
            iso_end_date = convert_to_iso_date(end_date)
        except ValueError:
            flash('Невалиден формат на датата (използвайте ДД-ММ-ГГГГ)', 'error')
            return redirect(url_for('reports'))
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        if ws is None:
            # This should never happen with openpyxl, but handle it for type safety
            flash('Error creating Excel workbook', 'error')
            return redirect(url_for('reports'))
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set title
        ws.title = "Detailed Measurements Report"
        
        # Query data
        query = '''
        SELECT p.product_name, d.dimension_name, ROUND(m.measured_value, 3) as measured_value, 
               d.nominal_value, d.tolerance_plus, d.tolerance_minus,
               m.measurement_date, m.inspector, m.machine_number, m.count, m.shift
        FROM measurements m
        JOIN dimensions d ON m.dimension_id = d.id
        JOIN products p ON m.product_id = p.id
        WHERE date(m.measurement_date) BETWEEN ? AND ?
        {}
        ORDER BY p.product_name, d.dimension_name, m.measurement_date DESC
        '''
        conditions = []
        params = [iso_start_date, iso_end_date]
        if product_id:
            conditions.append("m.product_id = ?")
            params.append(product_id)
        query = query.format("AND " + " AND ".join(conditions) if conditions else "")
        
        cursor.execute(query, params)
        report_data = [dict(row) for row in cursor.fetchall()]
        
        # Convert dates back to local format
        for row in report_data:
            row['measurement_date'] = convert_to_local_date(row['measurement_date'])
        
        # Set headers
        headers = ["Product", "Dimension", "Measured Value", "Nominal", "Tolerance (+/-)", "Measurement Date", "Inspector", "Machine", "Count", "Shift"]
        
        # --- Custom Excel Header ---
        from openpyxl.utils import get_column_letter
        last_col_letter = get_column_letter(len(headers))
        # Place the logo image in the left corner (A1) and match header height
        logo_path = os.path.join('static', 'images', 'nav-logo-impuls.png')
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.height = 60  # Match header height
            img.width = 120  # Proportional width
            ws.add_image(img, 'A1')
        # Merge B1 to last_col_letter1 for company name
        ws.merge_cells(f'B1:{last_col_letter}1')
        ws['B1'] = 'Impuls 97 Ltd'
        ws['B1'].font = Font(size=20, bold=True, color="FFFFFF")
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        # Subtitle
        ws.merge_cells(f'B2:{last_col_letter}2')
        ws['B2'] = 'Quality Control Report'
        ws['B2'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
        # Export date/time
        ws.merge_cells(f'B3:{last_col_letter}3')
        from datetime import datetime as dt
        ws['B3'] = f"Exported: {dt.now().strftime('%d-%m-%Y %H:%M')}"
        ws['B3'].font = Font(size=10, italic=True, color="FFFFFF")
        ws['B3'].alignment = Alignment(horizontal="right", vertical="center")
        # Fill blue background for header rows (A1 to last_col_letter3)
        for row in range(1, 4):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        # --- End Custom Header ---
        # Set headers (start from row 5)
        header_row = 5
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data (start from row 6)
        for row, data in enumerate(report_data, header_row + 1):
            ws.cell(row=row, column=1, value=data['product_name']).border = border
            ws.cell(row=row, column=2, value=data['dimension_name']).border = border
            ws.cell(row=row, column=3, value=data['measured_value']).border = border
            ws.cell(row=row, column=4, value=data['nominal_value']).border = border
            ws.cell(row=row, column=5, value=f"+{data['tolerance_plus']}/-{data['tolerance_minus']}").border = border
            ws.cell(row=row, column=6, value=data['measurement_date']).border = border
            ws.cell(row=row, column=7, value=data['inspector']).border = border
            ws.cell(row=row, column=8, value=data['machine_number']).border = border
            ws.cell(row=row, column=9, value=data['count']).border = border
            ws.cell(row=row, column=10, value=data.get('shift', '')).border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            # Safely retrieve the column index; fallback to 1 if it is None (prevents type checker errors)
            col_idx_safe = column[0].column if column[0].column is not None else 1
            column_letter = get_column_letter(col_idx_safe)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Set max width for 'Dimension' column (column 2)
            if column_letter == 'B':
                adjusted_width = min(max_length + 2, 20)  # Max 20 characters
            else:
                adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        filename = f"detailed_measurements_{start_date}_to_{end_date}.xlsx"
        
        # Save to memory
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

@app.route('/change_password', methods=['GET', 'POST'])
def change_password():
    """Route for admin users only - redirect others to dashboard"""
    if session.get('role') != 'admin':
        flash('Достъпът до тази страница е ограничен до администратори', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        current_user = session.get('user')
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']
        
        if not new_password or len(new_password) < 6:
            flash('Новата парола трябва да съдържа поне 6 символа', 'error')
        elif new_password != confirm_password:
            flash('Новата парола и потвърждението не съвпадат', 'error')
        else:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("UPDATE users SET password_hash=? WHERE username=?", 
                             (hash_password(new_password), current_user))
                conn.commit()
                flash('Паролата ви е променена успешно', 'success')
                return redirect(url_for('users'))
    
    return render_template('change_password.html', role=session.get('role'))

@app.route('/users', methods=['GET', 'POST'])
def users():
    if session.get('role') != 'admin':
        flash('Необходим е администраторски достъп', 'error')
        return redirect(url_for('dashboard'))
    with get_db_connection() as conn:
        cursor = conn.cursor()
        if request.method == 'POST':
            action = request.form.get('action')
            if action == 'add_user':
                username = request.form['username']
                password = request.form['password']
                role = request.form['role']
                if not username or not password:
                    flash('Потребителското име и паролата са задължителни', 'error')
                else:
                    try:
                        cursor.execute(
                            "INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
                            (username, hash_password(password), role)
                        )
                        conn.commit()
                        flash('Потребителят е добавен успешно', 'success')
                    except sqlite3.IntegrityError:
                        flash('Потребителското име вече съществува', 'error')
            elif action == 'delete_user':
                username = request.form['username']
                if username == session['user']:
                    flash('Не можете да изтриете собствения си акаунт', 'error')
                else:
                    cursor.execute("DELETE FROM users WHERE username=?", (username,))
                    conn.commit()
                    flash('Потребителят е изтрит успешно', 'success')
            elif action == 'change_password':
                username = request.form['username']
                new_password = request.form['new_password']
                confirm_password = request.form['confirm_password']
                
                if not new_password or len(new_password) < 6:
                    flash('Новата парола трябва да съдържа поне 6 символа', 'error')
                elif new_password != confirm_password:
                    flash('Новата парола и потвърждението не съвпадат', 'error')
                else:
                    cursor.execute("UPDATE users SET password_hash=? WHERE username=?", 
                                 (hash_password(new_password), username))
                    conn.commit()
                    flash(f'Паролата на потребител "{username}" е променена успешно', 'success')
        cursor.execute("SELECT username, role FROM users WHERE username != ?", (session['user'],))
        users = [dict(row) for row in cursor.fetchall()]
    return render_template('users.html', users=users, role=session.get('role'))

@app.route('/tolerance_tables', methods=['GET', 'POST'])
def tolerance_tables():
    # Check if user is admin
    if session.get('role') != 'admin':
        flash('Достъп отказан. Необходими са администраторски права.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get pagination parameters
    page = request.args.get('page', 1, type=int)
    per_page = 50  # Files per page
    
    # Directory for tolerance tables
    tolerance_dir = os.path.join('static', 'tolerance_tables')
    os.makedirs(tolerance_dir, exist_ok=True)
    
    # Get all PDF files in the tolerance_tables directory
    pdf_files = []
    if os.path.exists(tolerance_dir):
        pdf_files = [f for f in os.listdir(tolerance_dir) if f.lower().endswith('.pdf')]
        pdf_files.sort()  # Sort alphabetically
    
    # Calculate pagination
    total = len(pdf_files)
    total_pages = (total + per_page - 1) // per_page
    has_prev = page > 1
    has_next = page < total_pages
    
    # Get files for current page
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paginated_files = pdf_files[start_idx:end_idx]
    
    # Handle file upload
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'upload_tolerance':
            if 'tolerance_file' not in request.files:
                flash('Няма избран файл', 'error')
            else:
                file = request.files['tolerance_file']
                if file.filename == '':
                    flash('Няма избран файл', 'error')
                elif not file.filename.lower().endswith('.pdf'):
                    flash('Позволени са само PDF файлове', 'error')
                else:
                    # Check file size (max 50MB for tolerance tables)
                    file.seek(0, 2)
                    file_size = file.tell()
                    file.seek(0)
                    if file_size > 50 * 1024 * 1024:
                        flash('Размерът на файла трябва да е по-малък от 50MB', 'error')
                    else:
                        # Generate safe filename
                        filename = secure_filename(file.filename)
                        # Add timestamp to avoid conflicts
                        name, ext = os.path.splitext(filename)
                        timestamp = get_bulgarian_time().strftime('%Y%m%d_%H%M%S')
                        safe_filename = f"{name}_{timestamp}{ext}"
                        
                        file_path = os.path.join(tolerance_dir, safe_filename)
                        try:
                            file.save(file_path)
                            flash(f'Таблицата с допуски "{filename}" е качена успешно', 'success')
                            logger.info(f"Tolerance table uploaded: {safe_filename}")
                        except Exception as e:
                            flash('Грешка при запис на файла', 'error')
                            logger.error(f"Error uploading tolerance table: {e}")
        
        elif action == 'delete_tolerance':
            filename = request.form.get('filename')
            if filename:
                file_path = os.path.join(tolerance_dir, filename)
                if os.path.exists(file_path):
                    try:
                        os.remove(file_path)
                        flash(f'Таблицата с допуски "{filename}" е изтрита успешно', 'success')
                        logger.info(f"Tolerance table deleted: {filename}")
                    except Exception as e:
                        flash('Грешка при изтриване на файла', 'error')
                        logger.error(f"Error deleting tolerance table: {e}")
                else:
                    flash('Файлът не съществува', 'error')
        
        # Redirect to avoid form resubmission
        return redirect(url_for('tolerance_tables', page=page))
    
    return render_template('tolerance_tables.html',
                         files=paginated_files,
                         role=session.get('role'),
                         page=page,
                         total_pages=total_pages,
                         has_prev=has_prev,
                         has_next=has_next,
                         total=total)

@app.route('/view_tolerance_table/<filename>')
def view_tolerance_table(filename):
    tolerance_dir = os.path.join('static', 'tolerance_tables')
    file_path = os.path.join(tolerance_dir, filename)
    
    if not os.path.exists(file_path):
        flash('Файлът не съществува', 'error')
        return redirect(url_for('tolerance_tables'))
    
    try:
        return send_file(file_path, as_attachment=False, mimetype='application/pdf')
    except Exception as e:
        logger.error(f"Error serving tolerance table: {e}")
        flash('Грешка при зареждане на файла', 'error')
        return redirect(url_for('tolerance_tables'))

@app.route('/upload_drawing', methods=['POST'])
def upload_drawing():
    logger.info("Upload drawing request received")
    # Check if user is admin
    if session.get('role') != 'admin':
        logger.warning("Upload drawing: Access denied - not admin")
        response = jsonify({'status': 'error', 'message': 'Access denied. Admin rights required.'})
        response.headers['Content-Type'] = 'application/json'
        return response
    
    # Check if file was uploaded
    if 'drawing_file' not in request.files:
        response = jsonify({'status': 'error', 'message': 'No file selected'})
        response.headers['Content-Type'] = 'application/json'
        return response
    
    file = request.files['drawing_file']
    if not file or not hasattr(file, 'filename') or file.filename is None or file.filename == '':
        response = jsonify({'status': 'error', 'message': 'No file selected'})
        response.headers['Content-Type'] = 'application/json'
        return response
    
    if not str(file.filename).lower().endswith('.pdf'):
        response = jsonify({'status': 'error', 'message': 'Only PDF files are allowed'})
        response.headers['Content-Type'] = 'application/json'
        return response
    
    # Check file size without consuming the file stream
    file.seek(0, 2)  # Seek to end
    file_size = file.tell()
    file.seek(0)  # Reset to beginning
    if file_size > 10 * 1024 * 1024:
        response = jsonify({'status': 'error', 'message': 'File size must be less than 10MB'})
        response.headers['Content-Type'] = 'application/json'
        return response
    original_name = os.path.splitext(secure_filename(str(file.filename)))[0]
    unique_filename = f"{original_name}_{uuid.uuid4().hex[:8]}.pdf"
    drawings_dir = os.path.join('static', 'drawings')
    os.makedirs(drawings_dir, exist_ok=True)
    file_path = os.path.join(drawings_dir, unique_filename)
    try:
        file.save(file_path)
        logger.info(f"Drawing uploaded successfully: {unique_filename}")
        response_data = {
            'status': 'success', 
            'message': 'Drawing uploaded successfully',
            'filename': unique_filename
        }
        logger.info(f"Sending response: {response_data}")
        response = jsonify(response_data)
        response.headers['Content-Type'] = 'application/json'
        return response
    except Exception as e:
        logger.error(f"Error saving file: {e}")
        error_response = {'status': 'error', 'message': 'Error saving file'}
        logger.info(f"Sending error response: {error_response}")
        response = jsonify(error_response)
        response.headers['Content-Type'] = 'application/json'
        return response

@app.route('/delete_drawing', methods=['POST'])
def delete_drawing():
    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Достъп отказан. Необходими са администраторски права.'})
    product_id = request.form['product_id']
    drawing_path = request.form['drawing_path']
    drawing_slot = int(request.form.get('drawing_slot', 1))
    if not drawing_path:
        return jsonify({'status': 'error', 'message': 'Няма предоставен път до чертеж'})
    file_path = os.path.join('static', 'drawings', os.path.basename(drawing_path))
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
        # Clear the correct drawing_path column in database
        with get_db_connection() as conn:
            cursor = conn.cursor()
            if drawing_slot == 2:
                cursor.execute("UPDATE products SET drawing_path_2=NULL WHERE id=?", (product_id,))
            else:
                cursor.execute("UPDATE products SET drawing_path=NULL WHERE id=?", (product_id,))
            conn.commit()
        return jsonify({'status': 'success', 'message': 'Чертежът е изтрит', 'slot': drawing_slot})
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Неуспешно изтриване на чертеж: {e}', 'slot': drawing_slot})

@app.route('/replace_drawing', methods=['POST'])
def replace_drawing():
    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Достъп отказан. Необходими са администраторски права.'})
    product_id = request.form['product_id']
    drawing_slot = int(request.form.get('drawing_slot', 1))
    if 'drawing_file' not in request.files:
        return jsonify({'status': 'error', 'message': 'Няма избран файл'})
    file = request.files['drawing_file']
    if not file or not hasattr(file, 'filename') or file.filename is None or file.filename == '':
        return jsonify({'status': 'error', 'message': 'Няма избран файл'})
    if not str(file.filename).lower().endswith('.pdf'):
        return jsonify({'status': 'error', 'message': 'Позволени са само PDF файлове'})
    file.seek(0, 2)
    file_size = file.tell()
    file.seek(0)
    if file_size > 10 * 1024 * 1024:
        return jsonify({'status': 'error', 'message': 'Размерът на файла трябва да е по-малък от 10MB'})
    original_name = os.path.splitext(secure_filename(str(file.filename)))[0]
    unique_filename = f"{original_name}_{uuid.uuid4().hex[:8]}.pdf"
    drawings_dir = os.path.join('static', 'drawings')
    os.makedirs(drawings_dir, exist_ok=True)
    file_path = os.path.join(drawings_dir, unique_filename)
    try:
        file.save(file_path)
        # Update the correct drawing_path column in database
        with get_db_connection() as conn:
            cursor = conn.cursor()
            if drawing_slot == 2:
                cursor.execute("UPDATE products SET drawing_path_2=? WHERE id=?", (f'drawings/{unique_filename}', product_id))
            else:
                cursor.execute("UPDATE products SET drawing_path=? WHERE id=?", (f'drawings/{unique_filename}', product_id))
            conn.commit()
        return jsonify({'status': 'success', 'message': 'Чертежът е заменен', 'filename': unique_filename, 'slot': drawing_slot})
    except Exception as e:
        logger.error(f"Error replacing drawing: {e}")
        return jsonify({'status': 'error', 'message': 'Грешка при запис на файла', 'slot': drawing_slot})

@app.route('/add_drawing_to_product', methods=['POST'])
def add_drawing_to_product():
    # Check if user is admin
    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Достъп отказан. Необходими са администраторски права.'})
    
    product_id = request.form.get('product_id')
    drawing_slot = int(request.form.get('drawing_slot', 1))
    
    if not product_id:
        return jsonify({'status': 'error', 'message': 'Липсва ID на продукта'})
    
    # Check if file was uploaded
    if 'drawing_file' not in request.files:
        return jsonify({'status': 'error', 'message': 'Няма избран файл'})
    
    file = request.files['drawing_file']
    if not file or not hasattr(file, 'filename') or file.filename is None or file.filename == '':
        return jsonify({'status': 'error', 'message': 'Няма избран файл'})
    
    if not str(file.filename).lower().endswith('.pdf'):
        return jsonify({'status': 'error', 'message': 'Позволени са само PDF файлове'})
    
    # Check file size
    file.seek(0, 2)
    file_size = file.tell()
    file.seek(0)
    if file_size > 10 * 1024 * 1024:
        return jsonify({'status': 'error', 'message': 'Размерът на файла трябва да е по-малък от 10MB'})
    
    # Generate unique filename
    original_name = os.path.splitext(secure_filename(str(file.filename)))[0]
    unique_filename = f"{original_name}_{uuid.uuid4().hex[:8]}.pdf"
    drawings_dir = os.path.join('static', 'drawings')
    os.makedirs(drawings_dir, exist_ok=True)
    file_path = os.path.join(drawings_dir, unique_filename)
    
    try:
        # Save the file
        file.save(file_path)
        
        # Update the database
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Check if the product exists
            cursor.execute("SELECT id FROM products WHERE id=?", (product_id,))
            if not cursor.fetchone():
                # Clean up uploaded file if product doesn't exist
                if os.path.exists(file_path):
                    os.remove(file_path)
                return jsonify({'status': 'error', 'message': 'Продуктът не съществува'})
            
            # Update the appropriate drawing slot
            if drawing_slot == 2:
                cursor.execute("UPDATE products SET drawing_path_2=? WHERE id=?", (f'drawings/{unique_filename}', product_id))
            else:
                cursor.execute("UPDATE products SET drawing_path=? WHERE id=?", (f'drawings/{unique_filename}', product_id))
            
            conn.commit()
        
        logger.info(f"Drawing added to product {product_id}: {unique_filename}")
        return jsonify({
            'status': 'success', 
            'message': 'Чертежът е добавен успешно',
            'filename': unique_filename,
            'slot': drawing_slot
        })
    
    except Exception as e:
        logger.error(f"Error adding drawing to product: {e}")
        # Clean up uploaded file if database operation fails
        if os.path.exists(file_path):
            os.remove(file_path)
        return jsonify({'status': 'error', 'message': 'Грешка при добавяне на чертежа'})

@app.route('/recent_measurements')
def recent_measurements():
    with get_db_connection() as conn:
        cursor = conn.cursor()
        query = '''
        SELECT p.product_name, d.dimension_name, m.measured_value, d.nominal_value, d.tolerance_plus, d.tolerance_minus,
               m.measurement_date, m.inspector, m.machine_number, m.count, m.shift
        FROM measurements m
        JOIN dimensions d ON m.dimension_id = d.id
        JOIN products p ON m.product_id = p.id
        ORDER BY m.measurement_date DESC
        LIMIT 10
        '''
        cursor.execute(query)
        report_data = [dict(row) for row in cursor.fetchall()]
        # Convert dates to local format and add in_tolerance
        for row in report_data:
            row['measurement_date'] = convert_to_local_date(row['measurement_date'])
            nominal = row['nominal_value']
            minus = row['tolerance_minus']
            plus = row['tolerance_plus']
            measured = row['measured_value']
            row['in_tolerance'] = (nominal - minus) <= measured <= (nominal + plus)
        headers = [
            "Product", "Dimension", "Measured Value", "Nominal", "Tolerance (+/-)",
            "Measurement Date", "Inspector", "Machine", "Count", "Shift", "Tol. check"
        ]
    return render_template('recent_measurements.html', report_data=report_data, headers=headers, role=session.get('role'))

@app.route('/molds')
def molds():
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
        SELECT m.id, m.mold_name, m.mold_number, m.total_cycles, m.maintenance_threshold, 
               m.last_maintenance_date, m.status, m.created_date, p.product_name
        FROM molds m
        JOIN products p ON m.product_id = p.id
        ORDER BY m.created_date DESC
        ''')
        molds = [dict(row) for row in cursor.fetchall()]
        
        # Calculate maintenance status for each mold
        for mold in molds:
            cycles_remaining = mold['maintenance_threshold'] - mold['total_cycles']
            if cycles_remaining <= 0:
                mold['maintenance_status'] = 'overdue'
                mold['status_color'] = 'red'
            elif cycles_remaining <= 5000:
                mold['maintenance_status'] = 'due_soon'
                mold['status_color'] = 'orange'
            else:
                mold['maintenance_status'] = 'ok'
                mold['status_color'] = 'green'
    
    return render_template('molds.html', molds=molds, role=session.get('role'))

@app.route('/mold/<int:mold_id>')
def mold_detail(mold_id):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        
        # Get mold details
        cursor.execute('''
        SELECT m.*, p.product_name
        FROM molds m
        JOIN products p ON m.product_id = p.id
        WHERE m.id = ?
        ''', (mold_id,))
        mold = cursor.fetchone()
        
        if not mold:
            flash('Матрицата не е намерена', 'error')
            return redirect(url_for('molds'))
        
        # Get rework history
        cursor.execute('''
        SELECT * FROM rework_history
        WHERE mold_id = ?
        ORDER BY rework_date DESC
        ''', (mold_id,))
        rework_history = [dict(row) for row in cursor.fetchall()]
        
        # Get maintenance schedule
        cursor.execute('''
        SELECT * FROM maintenance_schedule
        WHERE mold_id = ?
        ORDER BY scheduled_date DESC
        ''', (mold_id,))
        maintenance_schedule = [dict(row) for row in cursor.fetchall()]
        
        # Get mold problems
        cursor.execute('''
        SELECT * FROM mold_problems
        WHERE mold_id = ?
        ORDER BY report_date DESC
        ''', (mold_id,))
        mold_problems = [dict(row) for row in cursor.fetchall()]
        
        # Convert dates to local format
        for item in rework_history:
            item['rework_date'] = convert_to_local_date(item['rework_date'])
        
        for item in maintenance_schedule:
            item['scheduled_date'] = convert_to_local_date(item['scheduled_date'])
            if item['completed_date']:
                item['completed_date'] = convert_to_local_date(item['completed_date'])
        
        for item in mold_problems:
            item['report_date'] = convert_to_local_date(item['report_date'])
    
    return render_template('mold_detail.html', 
                         mold=dict(mold), 
                         rework_history=rework_history, 
                         maintenance_schedule=maintenance_schedule,
                         mold_problems=mold_problems,
                         role=session.get('role'),
                         current_user=session.get('user'))

@app.route('/add_rework', methods=['POST'])
def add_rework():
    if not session.get('user'):
        return jsonify({'status': 'error', 'message': 'Необходимо е влизане в системата'})
    
    mold_id = request.form['mold_id']
    rework_type = request.form['rework_type']
    technician = request.form['technician']
    description = request.form.get('description', '')
    parts_replaced = request.form.get('parts_replaced', '')
    cost = request.form.get('cost', 0)
    
    if not all([mold_id, rework_type, technician]):
        return jsonify({'status': 'error', 'message': 'Всички задължителни полета трябва да бъдат попълнени'})
    
    try:
        cost = float(cost) if cost else 0
    except ValueError:
        cost = 0
    
    current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    with get_db_connection() as conn:
        cursor = conn.cursor()
        try:
            cursor.execute('''
            INSERT INTO rework_history (mold_id, rework_type, rework_date, technician, description, parts_replaced, cost)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (mold_id, rework_type, current_date, technician, description, parts_replaced, cost))
            
            # Reset total cycles to 0 after rework
            cursor.execute('''
            UPDATE molds 
            SET total_cycles = 0
            WHERE id = ?
            ''', (mold_id,))
            
            conn.commit()
            return jsonify({'status': 'success', 'message': 'Ремонтът е записан успешно'})
        except Exception as e:
            return jsonify({'status': 'error', 'message': f'Грешка при записване на ремонта: {e}'})

@app.route('/add_maintenance', methods=['POST'])
def add_maintenance():
    if not session.get('user'):
        return jsonify({'status': 'error', 'message': 'Необходимо е влизане в системата'})
    
    mold_id = request.form['mold_id']
    maintenance_type = request.form['maintenance_type']
    scheduled_date = request.form['scheduled_date']
    technician = request.form.get('technician', '')
    checklist_items = request.form.get('checklist_items', '')
    notes = request.form.get('notes', '')
    
    if not all([mold_id, maintenance_type, scheduled_date]):
        return jsonify({'status': 'error', 'message': 'Всички задължителни полета трябва да бъдат попълнени'})
    
    try:
        iso_date = convert_to_iso_date(scheduled_date)
    except ValueError:
        return jsonify({'status': 'error', 'message': 'Невалиден формат на датата'})
    
    with get_db_connection() as conn:
        cursor = conn.cursor()
        try:
            cursor.execute('''
            INSERT INTO maintenance_schedule (mold_id, maintenance_type, scheduled_date, technician, checklist_items, notes)
            VALUES (?, ?, ?, ?, ?, ?)
            ''', (mold_id, maintenance_type, iso_date, technician, checklist_items, notes))
            conn.commit()
            return jsonify({'status': 'success', 'message': 'Поддръжката е планирана успешно'})
        except Exception as e:
            return jsonify({'status': 'error', 'message': f'Грешка при планиране на поддръжката: {e}'})

@app.route('/complete_maintenance', methods=['POST'])
def complete_maintenance():
    if not session.get('user'):
        return jsonify({'status': 'error', 'message': 'Необходимо е влизане в системата'})
    
    maintenance_id = request.form['maintenance_id']
    technician = request.form.get('technician', '')
    notes = request.form.get('notes', '')
    
    current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    with get_db_connection() as conn:
        cursor = conn.cursor()
        try:
            # Update maintenance record
            cursor.execute('''
            UPDATE maintenance_schedule 
            SET completed_date = ?, technician = ?, technician_notes = ?, status = 'completed'
            WHERE id = ?
            ''', (current_date, technician, notes, maintenance_id))
            
            # Get mold_id for updating last_maintenance_date
            cursor.execute('SELECT mold_id FROM maintenance_schedule WHERE id = ?', (maintenance_id,))
            mold_id = cursor.fetchone()['mold_id']
            
            # Update mold's last maintenance date and reset total cycles
            cursor.execute('''
            UPDATE molds 
            SET last_maintenance_date = ?, total_cycles = 0
            WHERE id = ?
            ''', (current_date, mold_id))
            
            conn.commit()
            return jsonify({'status': 'success', 'message': 'Поддръжката е завършена успешно'})
        except Exception as e:
            return jsonify({'status': 'error', 'message': f'Грешка при завършване на поддръжката: {e}'})

@app.route('/update_mold_threshold', methods=['POST'])
def update_mold_threshold():
    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Достъп отказан. Необходими са администраторски права.'})
    
    mold_id = request.form['mold_id']
    # Retrieve threshold value from form, ensure it is provided
    threshold_value_str = request.form.get('new_threshold') or request.form.get('threshold')
    if threshold_value_str is None:
        return jsonify({'status': 'error', 'message': 'Липсва стойност за праг'})

    try:
        threshold = int(threshold_value_str)
        if threshold <= 0:
            raise ValueError
    except ValueError:
        return jsonify({'status': 'error', 'message': 'Прагът трябва да бъде положително цяло число'})
    
    with get_db_connection() as conn:
        cursor = conn.cursor()
        try:
            cursor.execute('''
            UPDATE molds 
            SET maintenance_threshold = ?
            WHERE id = ?
            ''', (threshold, mold_id))
            conn.commit()
            return jsonify({'status': 'success', 'message': 'Прагът за поддръжка е актуализиран успешно'})
        except Exception as e:
            return jsonify({'status': 'error', 'message': f'Грешка при актуализиране на прага: {e}'})

@app.route('/add_mold_problem', methods=['POST'])
def add_mold_problem():
    if not session.get('user'):
        return jsonify({'status': 'error', 'message': 'Необходимо е влизане в системата'})
    
    mold_id = request.form['mold_id']
    problem_type = request.form['problem_type']
    description = request.form.get('description', '')
    comments = request.form.get('comments', '')
    inspector = session.get('user')  # Current logged-in user
    
    if not all([mold_id, problem_type]):
        return jsonify({'status': 'error', 'message': 'Всички задължителни полета трябва да бъдат попълнени'})
    
    current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    with get_db_connection() as conn:
        cursor = conn.cursor()
        try:
            cursor.execute('''
            INSERT INTO mold_problems (mold_id, problem_type, description, inspector, report_date, comments)
            VALUES (?, ?, ?, ?, ?, ?)
            ''', (mold_id, problem_type, description, inspector, current_date, comments))
            conn.commit()
            return jsonify({'status': 'success', 'message': 'Проблемът е записан успешно'})
        except Exception as e:
            return jsonify({'status': 'error', 'message': f'Грешка при записване на проблема: {e}'})

@app.route('/edit_mold_problem', methods=['POST'])
def edit_mold_problem():
    if not session.get('user'):
        return jsonify({'status': 'error', 'message': 'Необходимо е влизане в системата'})
    
    problem_id = request.form['problem_id']
    problem_type = request.form['problem_type']
    description = request.form.get('description', '')
    comments = request.form.get('comments', '')
    current_user = session.get('user')
    current_role = session.get('role')
    
    if not all([problem_id, problem_type]):
        return jsonify({'status': 'error', 'message': 'Всички задължителни полета трябва да бъдат попълнени'})
    
    with get_db_connection() as conn:
        cursor = conn.cursor()
        
        # Check if user has permission to edit this problem
        cursor.execute('SELECT inspector FROM mold_problems WHERE id = ?', (problem_id,))
        problem = cursor.fetchone()
        
        if not problem:
            return jsonify({'status': 'error', 'message': 'Проблемът не е намерен'})
        
        # Only allow inspector who created the problem or admin to edit
        if problem['inspector'] != current_user and current_role != 'admin':
            return jsonify({'status': 'error', 'message': 'Нямате права за редактиране на този проблем'})
        
        try:
            cursor.execute('''
            UPDATE mold_problems 
            SET problem_type = ?, description = ?, comments = ?
            WHERE id = ?
            ''', (problem_type, description, comments, problem_id))
            conn.commit()
            return jsonify({'status': 'success', 'message': 'Проблемът е обновен успешно'})
        except Exception as e:
            return jsonify({'status': 'error', 'message': f'Грешка при обновяване на проблема: {e}'})

@app.route('/delete_mold_problem', methods=['POST'])
def delete_mold_problem():
    if not session.get('user'):
        return jsonify({'status': 'error', 'message': 'Необходимо е влизане в системата'})
    
    problem_id = request.form['problem_id']
    current_user = session.get('user')
    current_role = session.get('role')
    
    with get_db_connection() as conn:
        cursor = conn.cursor()
        
        # Check if user has permission to delete this problem
        cursor.execute('SELECT inspector FROM mold_problems WHERE id = ?', (problem_id,))
        problem = cursor.fetchone()
        
        if not problem:
            return jsonify({'status': 'error', 'message': 'Проблемът не е намерен'})
        
        # Only allow inspector who created the problem or admin to delete
        if problem['inspector'] != current_user and current_role != 'admin':
            return jsonify({'status': 'error', 'message': 'Нямате права за изтриване на този проблем'})
        
        try:
            cursor.execute('DELETE FROM mold_problems WHERE id = ?', (problem_id,))
            conn.commit()
            return jsonify({'status': 'success', 'message': 'Проблемът е изтрит успешно'})
        except Exception as e:
            return jsonify({'status': 'error', 'message': f'Грешка при изтриване на проблема: {e}'})

@app.route('/molds_dashboard')
def molds_dashboard():
    # Get pagination parameters only (keep existing filter system)
    page = request.args.get('page', 1, type=int)
    per_page = 50  # Molds per page
    
    with get_db_connection() as conn:
        cursor = conn.cursor()
        
        # Get total count for pagination
        cursor.execute("SELECT COUNT(DISTINCT m.id) FROM molds m")
        total = cursor.fetchone()[0]
        
        # Calculate pagination info
        total_pages = (total + per_page - 1) // per_page
        has_prev = page > 1
        has_next = page < total_pages
        offset = (page - 1) * per_page
        
        # Get molds with pagination
        cursor.execute('''
        SELECT m.id, m.mold_name, m.mold_number, m.total_cycles, m.maintenance_threshold, 
               m.last_maintenance_date, m.status, m.created_date, m.specifications_pdf, p.product_name,
               COUNT(mp.id) as problem_count
        FROM molds m
        JOIN products p ON m.product_id = p.id
        LEFT JOIN mold_problems mp ON m.id = mp.mold_id
        GROUP BY m.id, m.mold_name, m.mold_number, m.total_cycles, m.maintenance_threshold, 
                 m.last_maintenance_date, m.status, m.created_date, m.specifications_pdf, p.product_name
        ORDER BY m.created_date DESC
        LIMIT ? OFFSET ?
        ''', (per_page, offset))
        molds = [dict(row) for row in cursor.fetchall()]
        
        # Get current machine-mold assignments
        cursor.execute('''
        SELECT mma.machine_number, m.mold_name, m.mold_number, m.id as mold_id, p.product_name, 
               mma.assigned_date, mma.assigned_by, mma.status
        FROM machine_mold_assignments mma
        JOIN molds m ON mma.mold_id = m.id
        JOIN products p ON m.product_id = p.id
        WHERE mma.status = 'active'
        ORDER BY mma.assigned_date DESC
        ''')
        machine_assignments = [dict(row) for row in cursor.fetchall()]
        
        # Convert assignment dates to local format
        for assignment in machine_assignments:
            assignment['assigned_date'] = convert_to_local_date(assignment['assigned_date'])
        
        # Get recent machine activity from measurements
        cursor.execute('''
        SELECT mlp.machine_number, p.product_name, m.mold_name, m.mold_number, 
               mlp.last_update, mlp.last_count
        FROM machine_last_product mlp
        JOIN products p ON mlp.last_product_id = p.id
        LEFT JOIN molds m ON p.id = m.product_id
        WHERE mlp.last_update IS NOT NULL
        ORDER BY mlp.last_update DESC
        LIMIT 10
        ''')
        recent_activity = [dict(row) for row in cursor.fetchall()]
        
        # Convert activity dates to local format
        for activity in recent_activity:
            if activity['last_update']:
                activity['last_update'] = convert_to_local_date(activity['last_update'])
        
        # Get molds with recently added problems (max 5)
        cursor.execute('''
        SELECT m.id, m.mold_name, m.mold_number, p.product_name,
               mp.problem_type, mp.report_date, mp.inspector,
               mp.description
        FROM molds m
        JOIN products p ON m.product_id = p.id
        JOIN mold_problems mp ON m.id = mp.mold_id
        ORDER BY mp.report_date DESC
        LIMIT 5
        ''')
        recent_problems = [dict(row) for row in cursor.fetchall()]
        
        # Convert problem dates to local format
        for problem in recent_problems:
            problem['report_date'] = convert_to_local_date(problem['report_date'])
        
    return render_template('molds_dashboard.html', 
                         molds=molds, 
                         machine_assignments=machine_assignments,
                         recent_activity=recent_activity,
                         recent_problems=recent_problems,
                         role=session.get('role'),
                         page=page,
                         total_pages=total_pages,
                         has_prev=has_prev,
                         has_next=has_next,
                         total=total)

@app.route('/get_mold_problems/<int:mold_id>')
def get_mold_problems(mold_id):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        
        # Get mold info
        cursor.execute('''
        SELECT m.mold_name, m.mold_number, p.product_name
        FROM molds m
        JOIN products p ON m.product_id = p.id
        WHERE m.id = ?
        ''', (mold_id,))
        mold_info = cursor.fetchone()
        
        if not mold_info:
            return jsonify({'status': 'error', 'message': 'Матрицата не е намерена'})
        
        # Get problems for this mold
        cursor.execute('''
        SELECT * FROM mold_problems
        WHERE mold_id = ?
        ORDER BY report_date DESC
        ''', (mold_id,))
        problems = [dict(row) for row in cursor.fetchall()]
        
        # Convert dates to local format
        for problem in problems:
            problem['report_date'] = convert_to_local_date(problem['report_date'])
        
        return jsonify({
            'status': 'success',
            'mold_info': dict(mold_info),
            'problems': problems
        })

@app.route('/upload_mold_specifications', methods=['POST'])
def upload_mold_specifications():
    if not session.get('user'):
        return jsonify({'status': 'error', 'message': 'Необходимо е влизане в системата'})
    
    if 'pdf_file' not in request.files:
        return jsonify({'status': 'error', 'message': 'Не е избран файл'})
    
    file = request.files['pdf_file']
    mold_id = request.form.get('mold_id')
    
    if not file or file.filename == '':
        return jsonify({'status': 'error', 'message': 'Не е избран файл'})
    
    if not mold_id:
        return jsonify({'status': 'error', 'message': 'Не е указана матрица'})
    
    # Check if file is PDF
    if not file.filename or not file.filename.lower().endswith('.pdf'):
        return jsonify({'status': 'error', 'message': 'Само PDF файлове са разрешени'})
    
    # Generate unique filename
    filename = secure_filename(file.filename or '')
    name, ext = os.path.splitext(filename)
    unique_filename = f"mold_specs_{mold_id}_{int(time.time())}{ext}"
    
    # Save file
    file_path = os.path.join('static', 'drawings', unique_filename)
    file.save(file_path)
    
    # Update database
    with get_db_connection() as conn:
        cursor = conn.cursor()
        try:
            cursor.execute('UPDATE molds SET specifications_pdf = ? WHERE id = ?', 
                         (f'static/drawings/{unique_filename}', mold_id))
            conn.commit()
            return jsonify({'status': 'success', 'message': 'Спецификациите са качени успешно'})
        except Exception as e:
            # Delete file if database update fails
            if os.path.exists(file_path):
                os.remove(file_path)
            return jsonify({'status': 'error', 'message': f'Грешка при запазване: {e}'})

@app.route('/delete_mold_specifications', methods=['POST'])
def delete_mold_specifications():
    if not session.get('user'):
        return jsonify({'status': 'error', 'message': 'Необходимо е влизане в системата'})
    
    mold_id = request.form.get('mold_id')
    
    if not mold_id:
        return jsonify({'status': 'error', 'message': 'Не е указана матрица'})
    
    with get_db_connection() as conn:
        cursor = conn.cursor()
        
        # Get current PDF path
        cursor.execute('SELECT specifications_pdf FROM molds WHERE id = ?', (mold_id,))
        result = cursor.fetchone()
        
        if not result or not result['specifications_pdf']:
            return jsonify({'status': 'error', 'message': 'Няма качени спецификации за тази матрица'})
        
        pdf_path = result['specifications_pdf']
        file_path = pdf_path  # Path already contains 'static/drawings/'
        
        try:
            # Delete file
            if os.path.exists(file_path):
                os.remove(file_path)
            
            # Update database
            cursor.execute('UPDATE molds SET specifications_pdf = NULL WHERE id = ?', (mold_id,))
            conn.commit()
            
            return jsonify({'status': 'success', 'message': 'Спецификациите са изтрити успешно'})
        except Exception as e:
            return jsonify({'status': 'error', 'message': f'Грешка при изтриване: {e}'})

@app.route('/update_mold_comment', methods=['POST'])
def update_mold_comment():
    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Достъп отказан. Необходими са администраторски права.'})
    mold_id = request.form.get('mold_id')
    comment = request.form.get('comment', '').strip()
    if not mold_id:
        return jsonify({'status': 'error', 'message': 'Липсва ID на матрицата'})
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('UPDATE molds SET comment = ? WHERE id = ?', (comment, mold_id))
        conn.commit()
    return jsonify({'status': 'success', 'message': 'Коментарът е обновен успешно'})

@app.route('/delete_mold_comment', methods=['POST'])
def delete_mold_comment():
    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Достъп отказан. Необходими са администраторски права.'})
    mold_id = request.form.get('mold_id')
    if not mold_id:
        return jsonify({'status': 'error', 'message': 'Липсва ID на матрицата'})
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('UPDATE molds SET comment = NULL WHERE id = ?', (mold_id,))
        conn.commit()
    return jsonify({'status': 'success', 'message': 'Коментарът е изтрит успешно'})

@app.template_filter('bg_maintenance_type')
def bg_maintenance_type(value):
    mapping = {
        'preventive': 'Профилактична',
        'corrective': 'Корективна',
        'emergency': 'Аварийна',
        'scheduled': 'Планирана',
        'technical': 'Техническо обслужване',
        'polishing': 'Полиране',
        'part_replacement': 'Смяна на части',
        'cleaning': 'Почистване',
        'adjustment': 'Регулиране',
        'repair': 'Ремонт',
        'other': 'Друго',
    }
    return mapping.get(value, value)

@app.route('/complete_rework', methods=['POST'])
def complete_rework():
    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Достъп отказан. Необходими са администраторски права.'})
    rework_id = request.form.get('rework_id')
    if not rework_id:
        return jsonify({'status': 'error', 'message': 'Липсва ID на ремонта'})
    completed_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('UPDATE rework_history SET completed_date = ? WHERE id = ?', (completed_date, rework_id))
        conn.commit()
    return jsonify({'status': 'success', 'message': 'Ремонтът е отбелязан като завършен'})

if __name__ == '__main__':
    init_db()
    host = os.environ.get('FLASK_HOST', '127.0.0.1')
    port = int(os.environ.get('FLASK_PORT', 5000))
    app.run(host=host, port=port, debug=app.config['DEBUG'])